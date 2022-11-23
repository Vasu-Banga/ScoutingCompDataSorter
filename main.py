import openpyxl
import pygame

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("data.xlsx")

# Define variable to read sheet
dataframe1 = dataframe.active

teams = dataframe1['F']
print(teams[1].value)
pygame.init()
screen = pygame.display.set_mode((609,324))
pygame.display.set_caption("Auton Display")


img = pygame.image.load("./field.png").convert()

screen.blit(img, (0,0))

pygame.display.flip()
for row in range(dataframe1.max_row - 1):
    if(teams[row + 1].value == 111 ):
        # i = 0
        # for val in dataframe1[row + 2]:
        #     i += 1
        #     if i <= 24 and i >= 3 and i != 4:
        #         print(val.value)
        print("Row Number: " + str(row + 2))
        print("Auton Upper: " + str(dataframe1[row + 2][8].value))
        print("Auton Lower: " + str(dataframe1[row + 2][9].value))
        print("TeleOp Upper: " + str(dataframe1[row + 2][11].value))
        print("TeleOp Lower: " + str(dataframe1[row + 2][12].value))
        print("Was Defended: " + str(dataframe1[row + 2][13].value))
        print("Climbed To Rung: " + str(dataframe1[row + 2][17].value))
        print("Speed Rating: " + str(dataframe1[row + 2][21].value))
        print("Died Or Tipped: " + str(dataframe1[row + 2][22].value))
        if dataframe1[row + 2][4].value[0] == "r":
            if teams[row + 1].value == 1732:
                pygame.draw.rect(screen, (255,0,0), ((50.75 * int(dataframe1[row + 2][6].value % 12)),(54 * int(dataframe1[row + 2][6].value / 12)),50.75,54))
            else:
                pygame.draw.rect(screen, (255,0,255), ((50.75 * int(dataframe1[row + 2][6].value % 12)),(54 * int(dataframe1[row + 2][6].value / 12)),50.75,54))

        else:
            if teams[row + 1].value == 1732:
                pygame.draw.rect(screen, (0,0,255), ((50.75 * int(dataframe1[row + 2][6].value % 12)),(54 * int(dataframe1[row + 2][6].value / 12)),50.75,54))
            else:
                pygame.draw.rect(screen, (0,255,255), ((50.75 * int(dataframe1[row + 2][6].value % 12)),(54 * int(dataframe1[row + 2][6].value / 12)),50.75,54))
        print(dataframe1[row + 2][4].value[0])
        print("-----------------------------------------------------------------------")

# pygame.draw.rect(screen, (0,255,0), (0, 0,50.75,54)) #(r, g, b) is color, (x, y) is center, R is radius and w is the thickness of the circle border.
pygame.display.update()
#Step 45 per thingy, 1 is at (30,30)
#Step 55 in the y direction, 47 in the x direction

status = True
while status:
    for i in pygame.event.get():
 
        # if event object type is QUIT
        # then quitting the pygame
        # and program both.
        if i.type == pygame.QUIT:
            status = False

pygame.quit()                

























    # print(teams[row + 2].value)

# taxi = dataframe1['H']

# Iterate the loop to read the cell values
#24
# for col in dataframe1.iter_cols(min_row = 1, max_col = dataframe1.max_column, max_row = dataframe1.max_row, values_only=True):
#     print(col[0])
#     for row in range(dataframe1.max_row - 1):
#         print(col[row + 2])
    
