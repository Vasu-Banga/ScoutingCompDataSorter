import openpyxl

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("data.xlsx")

# Define variable to read sheet
database = dataframe.active

teams = database['F']

teamList = []
autonHigh = []
autonLow = []
teleOpHigh = []
teleOpLow = []
climb = []
speedRating = []
shootingPower = []
totalMatches = []
totalBrokenMatches = []

for row in range(database.max_row - 1):
    addData = True
    for value in teamList:
        if(teams[row + 1].value == value):
            addData = False
    if addData:
        teamList.append(teams[row + 1].value)
        autonHigh.append(0)
        autonLow.append(0)
        teleOpHigh.append(0)
        teleOpLow.append(0)
        climb.append(0)
        speedRating.append(0)
        shootingPower.append(0)
        totalMatches.append(0)
        totalBrokenMatches.append(0)
teamList.pop(-1)
autonHigh.pop(-1)
autonLow.pop(-1)
teleOpHigh.pop(-1)
teleOpLow.pop(-1)
climb.pop(-1)
speedRating.pop(-1)
shootingPower.pop(-1)
totalMatches.pop(-1)
totalBrokenMatches.pop(-1)

for row in range(database.max_row - 1):
    try:
        index = teamList.index(database[row + 2][5].value)
        if(database[row + 2][22].value == "N"):
            try:
                autonHigh[index] += database[row + 2][8].value
                autonLow[index] += database[row + 2][9].value
                teleOpHigh[index] += database[row + 2][11].value
                teleOpLow[index] += database[row + 2][12].value
            except:
                print("Error getting team ball values")    
            speedRating[index] += database[row + 2][21].value
            totalMatches[index] += 1
            if(database[row + 2][17].value != "x" and database[row + 2][17].value != "a"):
                climb[index] += database[row + 2][17].value
            else:
                climb[index] += 0
        else:
            totalBrokenMatches[index] += 1

    except:
        print("Error: Row " + str(row + 2))    

for val in range(len(teamList)):
    shootingPower[val] += autonHigh[val] * 4
    shootingPower[val] += autonLow[val] * 2
    shootingPower[val] += teleOpHigh[val] * 2
    shootingPower[val] += teleOpLow[val]



    

print("-------------------------------------------------")
print("Ranked teams based off of average shooting power")
tempRank = 1
chosenArr = shootingPower.copy()
#Comment this if you want totals instead of average
for val in range(len(chosenArr)):
    chosenArr[val] /= totalMatches[val]


rankedArr = chosenArr.copy()
rankedArr.sort(reverse=True)
tempTeamList = teamList.copy()

for val in rankedArr:
    try:
        index = chosenArr.index(val)
        print("Rank " + str(tempRank) + ": " + str(tempTeamList[index]) + "; {:0.2f}".format(chosenArr[index]))
        tempTeamList.pop(index)
        chosenArr.pop(index)
        tempRank += 1
    except:
        print("Error with value: " + str(val))
        print(rankedArr)
        print(chosenArr)
        print(tempTeamList)    

    

def findRank(arr, index):
    tempArr = arr.copy()
    for val in range(len(tempArr)):
        tempArr[val] /= totalMatches[val]
    tempArr.sort(reverse = True)
    # print(tempArr)
    # print(arr[index]/totalMatches[index])
    return (tempArr.index(arr[index]/totalMatches[index]) + 1)



while True:
    print("----------------------------------------------")
    print("What team do you want to learn about? ")
    queryTeam = input()
    queryTeam = int(queryTeam)
    try:
        
        index = teamList.index(queryTeam)
        try:
            autonHighRank = findRank(autonHigh, index)
            autonLowRank = findRank(autonLow, index)
            teleOpHighRank = findRank(teleOpHigh, index)
            teleOpLowRank = findRank(teleOpLow, index)
            climbRank = findRank(climb, index)
            speedRatingRank = findRank(speedRating, index)
            shootingPowerRank = findRank(shootingPower, index)
        except:
            print("Error getting ranks")    

        #Find rank of team in each category. May take a while, so feel free to comment out if unneeded
        













        try:
            print("Average High Auton: {:0.2f}".format(autonHigh[index]/totalMatches[index]) + ", ranked " + str(autonHighRank) + " in this category")
        except:
            print("Error in High Auton for team " + str(queryTeam))
        try:        
            print("Average Low Auton: {:0.2f}".format(autonLow[index]/totalMatches[index]) + ", ranked " + str(autonLowRank) + " in this category")
        except:
            print("Error in Low Auton for team " + str(queryTeam))
        try:        
            print("Average High TeleOp: {:0.2f}".format(teleOpHigh[index]/totalMatches[index]) + ", ranked " + str(teleOpHighRank) + " in this category")
        except:
            print("Error in High TeleOp for team " + str(queryTeam))
        try:        
            print("Average Low TeleOp: {:0.2f}".format(teleOpLow[index]/totalMatches[index]) + ", ranked " + str(teleOpLowRank) + " in this category")
        except:
            print("Error in Low TeleOp for team " + str(queryTeam))
        try:        
            print("Average Climb: {:0.2f}".format(climb[index]/totalMatches[index]) + ", ranked " + str(climbRank) + " in this category")
        except:
            print("Error in Climb for team " + str(queryTeam))
        try:    
            print("Average Speed Rating: {:0.2f}".format(speedRating[index]/totalMatches[index]) + ", ranked " + str(speedRatingRank) + " in this category")
        except:
            print("Error in Speed Rating for team " + str(queryTeam))
        try:        
            print("Average Shooting Power: {:0.2f}".format(shootingPower[index]/totalMatches[index]) + ", ranked " + str(shootingPowerRank) + " in this category")
        except:
            print("Error in Shooting Power for team " + str(queryTeam))
        try:    
            print("Data collected over " + str(totalMatches[index]) + " matches in which the robot was working")
        except:
            print("Error in getting total matches for team " + str(queryTeam))    
        try:
            print("Robot was incapacitated for " + str(totalBrokenMatches[index]) + " matches and could not run")   
        except:
            print("Error in getting total broken matches for team " + str(queryTeam))     
    except:
        print("That didn't work. Please try another team, or double check your syntax")
        print(teamList.index(111))

        